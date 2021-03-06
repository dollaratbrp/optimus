"""
This file manages all activities linked to Forecast

Author : Olivier Lefebvre

Update by : Nicolas Raymond
"""

# P2P Forecast for next 8 weeks

#  Used SQL Table and actions:
# [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS_EMAIL_ADDRESS] : Get data, alter, delete, send
# [Business_Planning].[dbo].[OTD_1_P2P_F_FORECAST_PARAMETERS] : Get data, alter, delete, send
# [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY] : Get data
# [Business_Planning].[dbo].[OTD_1_P2P_F_PRIORITY_WITHOUT_INVENTORY] : Get data
# OTD_1_P2P_D_INCLUDED_INVENTORY : Get data
# OTD_1_P2P_F_FORECAST_LOADS : Send data
# OTD_1_P2P_F_FORECAST_PRIORITY : Send data

from ParametersBox import *
from P2PFunctions import *
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from copy import deepcopy
import numpy as np


AutomaticRun = False  # set to True to automate code


def forecast():
    """
    Runs the forecast for the next 8 weeks

    """
    if not AutomaticRun:
        if not OpenParameters('FORECAST'):  # If user cancel request
            sys.exit()
        IsAdhoc = 1  # 1 for True, 0 for False (if automatic)
    else:
        IsAdhoc = 0  # 1 for True, 0 for False (if automatic)

    # -----------------------------------------------------------------------------------------------------------------#
    # -----------------------------------------------------------------------------------------------------------------#
    #                                             MODIFICATIONS SECTION
    # -----------------------------------------------------------------------------------------------------------------#
    # -----------------------------------------------------------------------------------------------------------------#

    # Path where the forecast results are saved
    saveFolder = 'S:\Shared\Business_Planning\Tool\Plant_to_plant\Optimus\P2P_excel_output\P2P_forecast\\'

    dayTodayComplete = pd.datetime.now().replace(second=0, microsecond=0)  # date to set in SQL for import date
    dayToday = weekdays(0)  # Date to display in report
    printLoads = False  # Print created loads
    drybox_sanity_check = True   # Activates drybox load validation
    good_credit_for_max = False  # If set to true, allow only wishes with good credit to be used to satisfy max
    save_log_file = False  # Save log file of process results
    wish_weekdays_range = 10  # Number of weekdays to consider from departure date when we filter wishlist
    result_time_stamp = time_now_string()
    general_folder = saveFolder + 'P2P_Forecast_' + dayToday + '\\'
    result_folder = general_folder + result_time_stamp + '\\'
    result_file = 'P2P_Forecast_'+result_time_stamp  # Name of excel file with today's date

    # -----------------------------------------------------------------------------------------------------------------#
    # -----------------------------------------------------------------------------------------------------------------#
    #                                            END OF MODIFICATIONS
    # -----------------------------------------------------------------------------------------------------------------#
    # -----------------------------------------------------------------------------------------------------------------#

    GeneralErrors = ''  # General errors met during execution

    ####################################################################################################################
    #                                                Get SQL DATA
    ####################################################################################################################

    # #If SQL queries crash
    downloaded = False
    nb_of_try = 0
    while not downloaded and nb_of_try < 3:  # 3 trials, SQL Queries sometime crash for no reason
        nb_of_try += 1
        try:
            downloaded = True

            ####################################################################################
            #                      Get details on point from and shipping point
            ####################################################################################
            shipping_points = shipping_point_names()

            ####################################################################################
            #                                 Email address Query
            ####################################################################################
            emails_list = get_emails_list('FORECAST')

            ####################################################################################
            #                                     Parameters Query
            ####################################################################################

            p2ps_list, connection = get_parameter_grid(forecast=True)

            ####################################################################################
            #                               Distinct date
            ####################################################################################

            #  We use the dates to loop through time
            inventory_connection = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_INVENTORY',
                                      headers='')
            date_query = """ SELECT distinct convert(date,[AVAILABLE_DATE]) as [AVAILABLE_DATE]
                      FROM [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY]
                      where [AVAILABLE_DATE] between convert(date,getdate()+1) and convert(date,getdate() + 7*8)
                      order by [AVAILABLE_DATE]
                    """

            dates = [single for sublist in inventory_connection.GetSQLData(date_query) for single in sublist]

            # We add tomorrow if it's not in the dates
            tomorrow = str(datetime.date.today() + datetime.timedelta(days=1))
            if tomorrow != dates[0]:
                dates.insert(0, tomorrow)

            ####################################################################################
            #                             Included Shipping_point
            ####################################################################################

            global DATAInclude
            get_nested_source_points(DATAInclude)

            ####################################################################################
            #                                Inventory and QA HOLD
            ####################################################################################

            inventory = get_inventory_and_qa()

            ####################################################################################
            #                                        Prod
            ####################################################################################
            # Filter based on if production is late, to make later
            prod_query = """ SELECT  [SHIPPING_POINT]
                                  ,[MATERIAL_NUMBER]
                                  ,[QUANTITY]
                                  ,convert(date,[AVAILABLE_DATE]) as AVAILABLE_DATE
                                  ,[STATUS]
                              FROM [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY]
                              where status = 'production PLAN'
                              and AVAILABLE_DATE between convert(date,getdate()+1) and convert(date, getdate() + 9*7)
                              order by AVAILABLE_DATE,SHIPPING_POINT,MATERIAL_NUMBER
                                    """

            production = [INVObj(*obj) for obj in inventory_connection.GetSQLData(prod_query)]

            ####################################################################################
            #                                 WishList Query
            ####################################################################################

            wishes = get_wish_list(forecast=True)

        except:
            downloaded = False
            print('SQL Query failed')

    # If SQL Queries failed
    if not downloaded:
        try:
            print('failed')
            send_email(emails_list, result_file, 'SQL QUERIES FAILED')
        except:
            pass
        sys.exit()

    # Creation of results folders
    create_directory(general_folder)
    create_directory(result_folder)

    # Creation of log file
    create_log_file(result_folder, save_log_file)

    # Application of credit rule to satisfy max
    set_good_credit_for_max(good_credit_for_max)

    ####################################################################################################################
    #                                     SQL tables declaration and cleaning
    ####################################################################################################################
    header = 'POINT_FROM,SHIPPING_POINT,QUANTITY,DATE,IMPORT_DATE,IS_ADHOC'
    loads_connection = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_FORECAST_LOADS', headers=header)
    loads_connection.deleteFromSQL()

    header = 'SALES_DOCUMENT_NUMBER,SALES_ITEM_NUMBER,SOLD_TO_NUMBER,POINT_FROM,SHIPPING_POINT,DIVISION,'\
                     'MATERIAL_NUMBER,SIZE_DIMENSIONS,LENGTH,WIDTH,HEIGHT,STACKABILITY,OVERHANG,QUANTITY,' \
                     'PRIORITY_RANK,X_IF_MANDATORY,PROJECTED_DATE,IMPORT_DATE,IS_ADHOC'
    priority_connection = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_FORECAST_PRIORITY',
                                headers=header)
    priority_connection.deleteFromSQL()

    ####################################################################################################################
    #                                           Excel Workbook declaration
    ####################################################################################################################

    # Initialization of workbook and filling option to warn user in output
    wb = Workbook()

    # Summary
    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    summary_columns = ['POINT_FROM', 'SHIPPING_POINT', 'QUANTITY', 'DEPARTURE_DATE']

    # Detailed
    detailed_ws = wb.create_sheet("DETAILED")
    detailed_columns = ['POINT_FROM', 'SHIPPING_POINT', 'DIVISION', 'MATERIAL_NUMBER',
                        'SIZE_DIMENSIONS', 'QUANTITY', 'DEPARTURE_DATE']

    columns_width = [20]*(len(detailed_columns)-1) + [25]
    worksheet_formatting(detailed_ws, detailed_columns, columns_width)

    ####################################################################################################################
    #                                    Set some LoadBuilder class' attributes
    ####################################################################################################################

    # We first set flatbed_48 as trailer reference of LoadBuilder for sanity check
    set_trailer_reference(get_trailers_data(['FLATBED_48'], [1]))

    # We set the attribute "validate_with_ref" for LoadBuilder class
    LoadBuilder.validate_with_ref = drybox_sanity_check

    ####################################################################################################################
    #                             Creation of the function to send wish assignment results
    ####################################################################################################################

    save_wish_assignment = build_forecast_output_sending_function(wishes, priority_connection,
                                                                  dayTodayComplete, detailed_ws)

    ####################################################################################################################
    #                                                Main loop for everyday
    ####################################################################################################################

    # We set INVObjs attribute Future to false to be able to use P2P normal process' functions
    for obj in inventory:
        obj.Future = False

    # We initialize a list of data that will be used to create the pivot table with pandas and a column counter
    summary_data = []
    column_counter = 3  # Starts at 3 because of the columns point from and point to

    # Initialization of the load bar
    progress = tqdm(total=len(dates), desc='Forecast progress')

    # START OF THE LOOP ################################################################################################

    for date in dates:

        # We shrink the wishlist to a 10 weekdays range
        filtered_wishes = [wish for wish in wishes if
                           wish.VALID_FROM_DATE <= weekdays(wish_weekdays_range, officialDay=date, return_as_date=True)
                           or wish.PERIOD_STATUS == 'P2P']

        # We write the departure date in the log file
        write_departure_date(date)

        # We reset the number of shared flatbed_53 and the residuals counter
        reset_flatbed_53()
        reset_residuals_counter()

        # we add today's prod and QA to inventory
        prod_used = []
        for i, item in enumerate(production):
            found = False
            if item.DATE > date:  # Ordered by date, so no need to continue
                break
            elif item.DATE == date and item.QUANTITY > 0:

                # We save is index to remove it later
                prod_used.append(i)

                for inv in inventory:
                    if inv.POINT == item.POINT and inv.MATERIAL_NUMBER == item.MATERIAL_NUMBER:
                        found = True
                        inv.QUANTITY += item.QUANTITY
                        item.QUANTITY = 0
                        break  # we found the good inv

                # We add the object if it wasn't already existing in the inventory
                if not found:
                    item.Future = False
                    inventory.append(item)

        # We remove prod used
        remove_indexes_from_list(production, prod_used)

        # Inventory is now updated.

        ################################################################################################################
        #                                 Perfect match and first loads creation
        ################################################################################################################

        # Now we create new loadBuilders
        for param in p2ps_list:
            param.reset()

        # Isolate perfect match
        approved_wishes = find_perfect_match(filtered_wishes, inventory, p2ps_list)

        # First loads creation
        perfect_match_loads_construction(p2ps_list, approved_wishes, print_loads=printLoads,
                                         assignment_function=save_wish_assignment, inventory_available_date=date)

        ################################################################################################################
        #                                       Satisfy minimums
        ################################################################################################################

        # Try to Make the minimum number of loads for each P2P
        satisfy_max_or_min(filtered_wishes, inventory, p2ps_list, print_loads=printLoads,
                           assignment_function=save_wish_assignment, inventory_available_date=date)

        ################################################################################################################
        #                                       Satisfy maximums
        ################################################################################################################

        satisfy_max_or_min(filtered_wishes, inventory, p2ps_list, print_loads=printLoads, satisfy_min=False,
                           assignment_function=save_wish_assignment, inventory_available_date=date)

        ################################################################################################################
        #                                     Leftover distribution
        ################################################################################################################

        satisfy_max_or_min(filtered_wishes, inventory, p2ps_list, print_loads=printLoads, leftovers=True,
                           assignment_function=save_wish_assignment, inventory_available_date=date)

        ################################################################################################################
        #                                         Results saving
        ################################################################################################################

        # We set a boolean to false that will indicate if there's anything that has been schedule this date
        consider_date_in_output = False

        # We save the loads created for that day
        for param in p2ps_list:

            # print('\nFROM :', param.POINT_FROM, 'TO :', param.POINT_TO)
            # print('DATE OF PLANIFICATION ',  weekdays(-1, officialDay=date))
            # print('DATE OF INVENTORY AVAILABILITY :', date)
            # print('PROJECTED DATE :', weekdays(max(0, param.days_to-1), officialDay=date), '\n')

            loads_connection.sendToSQL([(param.POINT_FROM, param.POINT_TO, len(param.LoadBuilder),
                                        weekdays(max(0, param.days_to-1), officialDay=date),
                                         dayTodayComplete, IsAdhoc)])

            if len(param.LoadBuilder) > 0:
                consider_date_in_output = True
                summary_data.append([shipping_points[param.POINT_FROM], shipping_points[param.POINT_TO],
                                     len(param.LoadBuilder), weekdays(max(0, param.days_to-1), officialDay=date)])

        # Update progress bar and column counter
        progress.update()
        column_counter += int(consider_date_in_output)

        # END OF THE LOOP ##############################################################################################

    # Loads were made for the next 8 weeks, now we send not assigned wishes to SQL
    for wish in wishes:
        if wish.QUANTITY != wish.ORIGINAL_QUANTITY:
            GeneralErrors += 'Error : this unit was not assigned and has a different quantity; \n {0}'.format(wish.lineToXlsx())
        priority_connection.sendToSQL(wish.lineToXlsx(dayTodayComplete))

    # We format the summary worksheet and save the workbook and the reference
    fake_columns = [''] * column_counter
    worksheet_formatting(summary_ws, fake_columns, [20] * len(fake_columns))
    reference = [savexlsxFile(wb, result_folder, result_file)]
    wb.close()

    # We create the pivot table and write it in the already saved xlsx file
    workbook = load_workbook(reference[0])
    writer = pd.ExcelWriter(reference[0], engine='openpyxl')
    summary_data = pd.DataFrame(data=summary_data, columns=summary_columns)
    summary_data = pd.pivot_table(summary_data, values='QUANTITY', index=['POINT_FROM', 'SHIPPING_POINT'],
                                  columns=['DEPARTURE_DATE'], aggfunc=np.sum)
    writer.book = workbook
    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    summary_data.to_excel(writer, sheet_name="SUMMARY", header=True)
    writer.save()

    # We send the emails
    send_email(emails_list, result_file, 'P2P Forecast is now updated.\n'+GeneralErrors, reference)

