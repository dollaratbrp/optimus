"""

Authors : Olivier Lefebvre
          Nicolas Raymond

This file creates loads for P2P

Last update : 2020-01-07
By : Nicolas Raymond

"""

import os
import openpyxl
import smtplib
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from email.mime.base import MIMEBase
import datetime
import pandas as pd
import pyodbc
from string import ascii_uppercase

timeReference = datetime.datetime.now()


def savexlsxFile(wb, path, filename, Time=False, extension='.xlsx'):

    """
    Saves xlsx files

    :param wb: workbook
    :param path: path where the file is going to be save (leave '' if no path)
    :param filename: Name of the document
    :param Time: (bool) True to write the creation time in the file name
    :param extension: extension of the file
    :return: complete path of the file saved
    """
    
    if Time is True:
        time_string = time_now_string()
    else:
        time_string = ''
    ErrorMessage = ''
    
    saved = False
    nameExtension = ''
    while not saved:
        try:
            saved = True
            wb.save(path+filename+nameExtension+time_string+extension)
        except:
            msg = "Someone already has the file "+filename+nameExtension+extension+" open"
            ErrorMessage += msg
            saved = False
            nameExtension += " Copy"
    return path+filename+nameExtension+time_string+extension


def send_email(recipient, subject, text, listfile=[], CC=[]):
    """
    Sends email

    :param recipient: list of email addresses that will receive the email
    :param subject: subject of the email
    :param text: text contained in the email
    :param listfile: list with directories of the files to send with email
    :param CC: list of email adresses in CC

    Example :
        send_email(["felix.theroux-joncas@brp.com"], "Subject", "text", ["C:/Users/test.txt", "C:/Users/test2.xlsx"])
    """
    email = "SVC_CAVL_SMTP_OTD"
    password = "!983IIdj111" 
    msg = MIMEMultipart()
    msg['From'] = "OrdertoDelivery@brp.com"
    msg['To'] = ", ".join(recipient)
    msg['CC'] = ", ".join(CC)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject
    msg.attach(MIMEText(text))

    # Attach all file in listfile
    for pathfile in listfile:
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(pathfile, "rb").read())
        encoders.encode_base64(part)
        if '/' in pathfile:
            file = pathfile.split("/")[-1]
        else:
            file = pathfile.split('\\')[-1]
        part.add_header('Content-Disposition', 'attachment', filename=file)
        msg.attach(part)  
    try: 
        smtpObj = smtplib.SMTP_SSL('smtp-cavl.brp.local', 465)
        smtpObj.ehlo()
        smtpObj.login(email, password)
        smtpObj.sendmail("OrdertoDelivery@brp.com", recipient+CC, msg.as_string())
        smtpObj.close()
    except:
        print("Sending email failed")


def time_now_string():
    """
    Returns time extension for file
    """
    time_now = pd.datetime.now()
    time_string = '_' + str(time_now.year) + '.' + str(time_now.month) + '.' + str(time_now.day) + '__' +\
                  str(time_now.hour) + '.' + str(time_now.minute) + '.' + str(time_now.second)
    return time_string


def timeSinceLastCall(functionName="", ToPrint=True):
    """
    Print the time since the last call, gives an indication of the time required by each function

    :param functionName: name of the last executed function
    :param ToPrint: (bool) True to print the time required
    """

    global timeReference
    if ToPrint:
        print(functionName+' : '+str((datetime.datetime.now()-timeReference).total_seconds())+" sec")
    timeReference = datetime.datetime.now()


def weekdays(dayAfter, startDay=0, officialDay='', return_as_date=False):
    """
    Returns the date in (dayAfter) week days from today's date

    :param dayAfter: Number of days from today
    :param startDay: First day to start from now (0 for today, -1 for yesterday, 1 for tomorrow,...)
    :param officialDay: date to consider as today in format '%Y-%m-%d'
    :param return_as_date: bool indicating if we return date in date format
    """

    if officialDay == '':
        today = datetime.date.today() + datetime.timedelta(startDay)
    else:
        today = datetime.datetime.strptime(officialDay, '%Y-%m-%d').date() + datetime.timedelta(startDay)
    wkday = today.weekday()

    # if today's date is a weekend day
    if wkday > 4:
        today = datetime.date.today()+datetime.timedelta(startDay+7-wkday)
    dayConversion = dayAfter % 5
    week = (dayAfter - dayConversion)/5
    date2 = today+datetime.timedelta(7*week + dayConversion)
    if date2.weekday() > 4 or date2.weekday() < dayConversion:
        date2 = date2+datetime.timedelta(2)
    deltmonth = ''  # else we have a problem with SQL date format
    delday=''
    if date2.month < 10:
        deltmonth = '0'
    if date2.day < 10:
        delday = '0'

    date_to_return = str(date2.year)+'-'+deltmonth+str(date2.month)+'-'+delday+str(date2.day)

    if return_as_date:
        return datetime.datetime.strptime(date_to_return, '%Y-%m-%d')
    else:
        return str(date2.year)+'-'+deltmonth+str(date2.month)+'-'+delday+str(date2.day)


def xlsxToTxtFile(readPath, filexlsx, writePath, txtfileName):
    """
    Writes excel data in a text document

    :param readPath: path where is the file we want to read
    :param filexlsx: name of the file with .xlsx extension included
    :param writePath: path where the new text file will be saved
    :param txtfileName: name of the new text file
    """

    global ErrorMessage
    wb = openpyxl.load_workbook(readPath+filexlsx)
    ws = wb.active
    try:
        txt_file = open(writePath+txtfileName, 'w')
        data = ""
        for rows in ws:
            line = ""
            for cells in rows: 
                line += str(cells.value)
                line += "\t"
            data += (line[:-1]+"\n")
        txt_file.write(data[:-1])
        txt_file.close()
    except:
        ErrorMessage += "Error in xlsxToTxtFile function"


def create_directory(directory):
    """
    Creates a directory if it doesn't already exists
    :param directory: directory to create
    """
    # Creation of results folder
    if not os.path.exists(directory):
        os.mkdir(directory)
        print("Directory ", directory, " Created ")
    else:
        print("Directory ", directory, " already exists")


def worksheet_formatting(ws, column_titles, column_widths, filling=None):
    """
    Does the formatting of an excel worksheet that will be use to store outputs
    :param ws: excel worksheet
    :param column_titles: list with all column titles
    :param column_widths: list with all column widths
    :param filling: PaternFill for the excel worksheet column titles
    """
    # We set the columns' title
    ws.append(column_titles)

    # We set the columns' width using alphabetical index
    alphabet = list(ascii_uppercase)
    for i in range(len(column_titles)):
        if i <= 25:
            ws.column_dimensions[alphabet[i]].width = column_widths[i]
        else:
            first_letter = alphabet[int(i / 26) - 1]
            second_letter = alphabet[(i % 26)]
            index = first_letter + second_letter
            ws.column_dimensions[index].width = column_widths[i]

    # We apply the filling to the columns' title
    if filling is None:
        filling = PatternFill(fill_type="solid", start_color="a6a6a6", end_color="a6a6a6")

    for i in range(1, len(column_titles)+1):
        ws.cell(row=1, column=i).fill = filling


def create_excel_table(ws, name, column_titles, TableStyle=None):
    """
    Creates an excel table with data from the worksheet
    :param ws: worksheet
    :param name: display name of the table
    :param column_titles: list with column titles
    :param TableStyle: Style to customize the table
    """
    nb_of_rows = get_number_of_rows(ws)
    if nb_of_rows > 1:
        table_reference = 'A1:' + get_last_column_index(column_titles) + str(nb_of_rows)
        tab = Table(displayName=name, ref=table_reference)
        ws.add_table(tab)


def get_last_column_index(column_titles):
    """
    Get the index letter associated to the last column of the worksheet
    :param column_titles: list with column titles
    :return : Uppercase letters of the alphabet associated with the last column of the worksheet
    """
    alphabet = list(ascii_uppercase)
    column_number = len(column_titles)
    if column_number <= 26:
        index = alphabet[column_number - 1]
    else:
        first_letter = alphabet[int(column_number/26)-1]
        second_letter = alphabet[(column_number % 26)-1]
        index = first_letter+second_letter

    return index


def get_number_of_rows(ws):
    """
    Gets the number of the rows used in a worksheet (based on first column length)
    :param ws: worksheet
    :return: number of rows
    """
    return len(ws['A'])


def apply_filling_to_rows(ws, row_indexes, nb_of_columns, filling=None):
    """
    Apply a filling to a whole row in a worksheet

    :param ws: worksheet
    :param row_indexes: list with indexes of the rows on which we want to apply the filling (start at 1 in excel)
    :param filling : PaternFill object from openpyxl.style
    :param nb_of_columns: number of columns in the worksheet
    """
    if filling is None:
        filling = PatternFill(fill_type="solid", start_color="a6a6a6", end_color="a6a6a6")

    for index in row_indexes:
        for j in range(1, nb_of_columns+1):
            ws.cell(row=index, column=j).fill = filling


def group_by_all_except_qty(dataframe_input, columns_name):
    """
    Create a data frame and execute a group by on all columns except QTY

    :param dataframe_input: list with lists of input to create pandas dataframe
    :param columns_name: list with columns name
    :return: pandas dataframe
    """
    frame = pd.DataFrame(data=dataframe_input, columns=columns_name)
    frame = frame.groupby([column for column in columns_name if column != 'QUANTITY']).sum()
    frame = frame.reset_index()

    return frame

class SQLConnection:

    """Connection to SQL servers"""

    def __init__(self, server, DataBase, table, headers=''):
        self.server = server
        self.DataBase = DataBase
        self.table = table
        self.conn = pyodbc.connect('Driver={SQL Server};'
                                   'Server='+server+';'
                                   'Database='+DataBase+';'
                                   'Trusted_Connection=yes;')
        self.cursor = cursor = self.conn.cursor()
        self.headers = headers

    def deleteFromSQL(self, conditions=''):
        """
        Deletes data from the sql table
        :param conditions: sql condition that indicates data to delete  (Example : NAMEOFCOLUMN < 0)
        """
    
        if conditions == '':
            application = ''
        else:
            application = ''' WHERE ''' + conditions
        query = ''' DELETE FROM ''' + self.table + application

        self.cursor.execute(query)
        self.conn.commit()

    def GetSQLData(self, SQLquery):
        """
        Gets data from the sql table
        :param SQLquery: sql query to get the data
        :return: list of lists containing the data
        """
        
        if SQLquery != """ """:
            self.cursor.execute(SQLquery)
            result = self.cursor.fetchall()
            resultList = [list(i) for i in result]
            return resultList
        else:  # Else returns an empty list
            return[]

    def UpdateSQL(self, RowsToChange, conditions):
        """
        Modifies data FROM a SQL table

        :param RowsToChange: Rows to change ( ex: LAST_CHANGED_DATE = GETDATE())
        :param conditions: Conditions to update DATA (ex: "country = 'CA' AND shipping_POINT <>'4110'")
        """
    
        query = ''' UPDATE '''+self.table+''' SET '''+RowsToChange+''' WHERE ''' + conditions

        self.cursor.execute(query)
        self.conn.commit()

    def sendToSQL(self, values, headers='', headersCount=-1):
        """

        :param values: list of lists of values
        :param headers:
        :param headersCount: number of columns in the table, it counts the number of comma in headers +1
        :return:
        """
        if headers == '':
            headers = self.headers
        if headersCount == -1:  
            headersCount = headers.count(',')
        valuesToInsert = '?'+headersCount*',?'   # expected number of values to insert (ex: 4 columns = '(?,?,?,?)')
    
        query = ''' INSERT INTO '''+self.table+''' ('''+headers+''')VALUES ('''+valuesToInsert+''')'''

        self.cursor.executemany(query, values)
        self.conn.commit()
