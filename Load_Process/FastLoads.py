"""
This file manages all activities linked to Optimus standalone mode

Author : Nicolas Raymond
"""
from tkinter import *
import pandas as pd
from openpyxl import load_workbook, Workbook
from InputOutput import SQLConnection, savexlsxFile, send_email
from P2PFunctions import get_trailers_data, get_emails_list
from LoadBuilder import LoadBuilder, set_trailer_reference
from random import randint
from InputOutput import worksheet_formatting, create_excel_table
from ParametersBox import change_emails_list, set_project_name, VerticalScrolledFrame
from datetime import datetime

plot_loads = True

workbook_path = 'U:\LoadAutomation\Optimus\FastLoadsSKUs.xlsx'

saving_path = 'U:\LoadAutomation\Optimus\\'


class FastLoadsBox(VerticalScrolledFrame):

    def __init__(self, master):

        """
        Generates a GUI that allows user to valid fast load details and to start the execution
        :param master: root of the tkinter window
        """

        # Saving of master and setting of GUI title
        self.master = master
        self.master.title('Optimus FastLoads')

        # Initializing the vertical scrolled frame
        super().__init__(self.master)
        self.pack()

        # Initialization of a bool indicating if we work with metal crates or not
        self.crate_type = StringVar()
        self.crate_type.set('W')

        # Initialization of a bool indicating if we want to do sanity check for drybox
        self.sanity_check = BooleanVar()

        # Initialization and positioning of a frame that will contain sku labels and entries
        self.sku_frame = LabelFrame(self.interior, borderwidth=2, relief=RIDGE, text='Crates')
        self.sku_frame.grid(row=0, column=0)

        # All single sku labels and qty entries initialization and positioning
        self.sku_labels, self.sku_entries = self.create_sku_labels_and_entries(self.sku_frame,
                                                                               self.read_skus_and_quantities())
        # Initialization of a frame for crate types radio buttons and initialization of the radio buttons themselves
        self.crate_type_frame = LabelFrame(self.sku_frame, text='Type')
        self.crate_type_frame.grid(row=len(self.sku_labels)+1, columnspan=2)
        self.metal_button = Radiobutton(self.crate_type_frame, text='Metal', variable=self.crate_type, value='M')
        self.wood_button = Radiobutton(self.crate_type_frame, text='Wood', variable=self.crate_type, value='W')
        self.metal_button.grid(row=0, column=0)
        self.wood_button.grid(row=0, column=1)

        # Recuperation of trailers data
        self.trailers_data = get_trailers_data()

        # Initialization and positioning of a frame that will contain trailers labels and entries
        self.trailer_frame = LabelFrame(self.interior, borderwidth=2, relief=RIDGE, text='Trailers')
        self.trailer_frame.grid(row=0, column=1)

        # All single trailer labels and qty entries initialization and positioning
        self.trailer_labels, self.trailer_qty_entries = self.create_trailer_labels_and_entries(self.trailer_frame)

        # Initialization and positioning of a frame for max loads entry and sanity check button
        self.max_frame = Frame(self.trailer_frame, borderwidth=2, relief=RIDGE)
        self.max_frame.grid(row=len(self.trailer_labels)+1, columnspan=2)
        self.max_label = Label(self.max_frame, text="MAX LOADS", padx=4, pady=10)
        self.max_entry = Entry(self.max_frame, justify='center', width=10)
        self.max_entry.insert(0, "∞")
        self.sc_radio_button = Checkbutton(self.max_frame, text='Drybox validation', onvalue=TRUE, offvalue=FALSE,
                                           variable=self.sanity_check)

        self.max_label.grid(row=0, column=0)
        self.max_entry.grid(row=0, column=1)
        self.sc_radio_button.grid(row=1, columnspan=2)

        # Modify emails list button configurations
        self.modify_email = Button(self.interior, text='Modify Emails List', padx=30, pady=10, bd=2,
                                   command=change_emails_list, font=('TkDefaultFont', 12, 'italic'), bg='gray80')
        self.modify_email.grid(row=2, column=0)

        # "Run optimus" button configurations
        self.run_button = Button(self.interior, text='Run Optimus', padx=30, pady=10, command=self.run_optimus, bd=2,
                                 font=('TkDefaultFont', 12, 'italic'), bg='gray80')
        self.run_button.grid(row=2, column=1, sticky=E+W)

        # Initialization of an empty load builder and an empty tracker attribute
        self.LoadBuilder = None
        self.tracker = None

    @staticmethod
    def create_sku_labels_and_entries(frame, skus_n_qty):
        """
        Creates and places wisely sku labels and qty entries in the frame entered as parameter
        :param frame: frame widget of our FastLoadsBox
        :param skus_n_qty: pandas dataframe containing SKUs and quantities associated to each
        :return: list of labels and list of entries
        """

        # Labels and entries container initialization
        labels = []
        entries = []

        for i in skus_n_qty.index:

            # Sku label initialization and positioning
            labels.append(Label(frame, text=str(skus_n_qty['SKU'][i]), padx=10, pady=10))
            labels[i].grid(row=i+1, column=0)

            # Qty entries initialization and positioning
            entries.append(Entry(frame, justify='center', width=10))
            entries[i].insert(0, str(skus_n_qty['QTY'][i]))
            entries[i].grid(row=i+1, column=1)

        return labels, entries

    def create_trailer_labels_and_entries(self, frame):
        """
        Creates and places wisely trailer labels and qty entries in the frame entered as parameter
        :param frame: frame widget of our FastLoadsBox
        :return: list of labels and list of entries
        """

        # Labels and entries container initialization
        labels = []
        entries = []

        for i in self.trailers_data.index:

            # Trailer label initialization and positioning
            labels.append(Label(frame, text=str(self.trailers_data['CATEGORY'][i]), padx=10, pady=10))
            labels[i].grid(row=i + 1, column=0)

            # Qty entries initialization and positioning
            entries.append(Entry(frame, justify='center', width=10))
            entries[i].insert(0, 0)
            entries[i].grid(row=i + 1, column=1)

        return labels, entries

    @staticmethod
    def read_skus_and_quantities():
        """
        Reads SKUs and quantities in the Excel table for fast loads
        :return: pandas dataframe containing SKUs and quantities associated to each
        """

        wb = load_workbook(workbook_path)
        ws = wb.active

        # We access to the table where the data on models is contained (table 0)
        table_range = ws._tables[0].ref

        # We save the data in a data frame
        skus_n_qty = build_dataframe(ws[table_range], fast_loads_input=True)

        return skus_n_qty

    def run_optimus(self):
        """
        Run Optimus program on our selection of SKUs
        :return:
        """
        # We save definitive quantities of SKUs and trailers
        skus_list, qty_list = self.save_skus_and_quantities()
        number_of_trailers = self.save_trailers_quantities()

        if len(skus_list) != 0 and number_of_trailers != 0:

            # Building of dataframes required for the loading
            # (df1) Use to build an object that keeps track of link between SKUs and size code
            # (df2) Dataframe needed by the load builder
            complete_dataframe = self.get_complete_dataframe(skus_list, qty_list, str(self.crate_type.get()))
            df1, df2 = self.split_dataframes(complete_dataframe)

            # Initialization of the tracker (size_code dictionaries with SKUsContainer as value)
            self.tracker = self.tracker_initialization(df1)

            # Initialization of our LoadBuilder
            self.LoadBuilder = LoadBuilder(trailers_data=self.trailers_data)

            # We set flatbed_48 as trailer reference of LoadBuilder for sanity check
            if self.sanity_check.get():
                set_trailer_reference(get_trailers_data(['FLATBED_48'], [1]))
                LoadBuilder.validate_with_ref = True

            # Recuperation of the maximum of loads
            max_loads = self.max_entry.get()
            if max_loads == '∞':
                max_loads = 5000
            else:
                max_loads = int(max_loads)

            # Closing of the GUI and construction of loads
            self.master.destroy()
            size_code_used = self.LoadBuilder.build(models_data=df2, max_load=max_loads, plot_load_done=plot_loads)

            # We save all the results in a workbook at the path mentioned
            reference = self.write_results(df2)

            # We send the email
            send_email(get_emails_list('ADHOC'), 'AdHoc ' + str(datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                       '', reference)

    def save_skus_and_quantities(self):
        """
        Save definitive lists of SKUs and quantities to use once user push "Run Optimus".
        :return: list of skus and list of qty
        """

        skus_list = []  # To find data in sql
        qty_list = []   # To use as column in pandas dataframe

        # We save each sku in a list if their quantity is greater than 0
        for i in range(len(self.sku_labels)):
            qty = int(self.sku_entries[i].get())
            if qty > 0:
                skus_list.append(self.sku_labels[i]['text'])
                qty_list.append(qty)

        return skus_list, qty_list

    def save_trailers_quantities(self):
        """
        Save definitive quantities of trailers in our trailer dataframe

        :return : total number of trailers
        """

        # Initialization of a counter of trailers
        number_of_trailers = 0

        for i in self.trailers_data.index:
            qty = int(self.trailer_qty_entries[i].get())
            self.trailers_data.at[i, 'QTY'] = qty
            number_of_trailers += qty

        return number_of_trailers

    def write_results(self, grouped_dataframe):
        """
        Saves all the results from the loading in an Excel workbook

        :param grouped_dataframe: pandas dataframe that was passed to our loadbuilder
        """

        # Initialization of a workbook
        wb = Workbook()

        # Approved loads worksheet construction and customization
        approved_ws = wb.active
        approved_ws.title = "APPROVED"
        columns_title = ['LOAD_NUMBER', 'CATEGORY', 'LOAD LENGTH', 'MATERIAL_NUMBER', 'QUANTITY', 'SIZE_DIMENSIONS']
        worksheet_formatting(approved_ws, columns_title, [20]*len(columns_title))

        # Unused crates worksheet construction and customization
        unused_ws = wb.create_sheet("UNUSED")
        unused_columns = ['MATERIAL_NUMBER', 'QUANTITY']
        worksheet_formatting(unused_ws, unused_columns, [20]*len(unused_columns))

        # Writing of approved loads
        self.write_approved_loads(approved_ws, len(columns_title), grouped_dataframe)

        # Writing of unused crates
        self.write_unused_crates(unused_ws, len(unused_columns))

        # Save the xlsx file
        return [savexlsxFile(wb=wb, path=saving_path, filename='AdHoc', Time=True)]

    def write_approved_loads(self, ws, nbr_of_cols, grouped_dataframe):
        """
        Writes the results for the approved loads in the worksheet passed as parameter

        :param ws: worksheet on which we write the results
        :param nbr_of_cols: number of columns used in the worksheet
        :param grouped_dataframe: pandas dataframe that was passed to our loadbuilder
        """

        # Retrieving summary of loads done
        loads = self.LoadBuilder.get_loading_summary()
        load_number = 0

        # We build the dataframe on which we'll do a groupby to simplify output
        data = []
        if len(self.LoadBuilder) > 0:  # If some loads were created

            # For every line of data in our "loads" dataframe
            for i in loads.index:

                # We save the trailer category and load length for this type of load
                category = loads['TRAILER'][i]
                load_length = loads['LOAD LENGTH'][i]

                # For every load of this kind
                for iteration in range(int(loads["QTY"][i])):

                    load_number += 1

                    # For all the different size_code column in the dataframe
                    for size_code in loads.columns[4::]:

                        # if there is some quantity for this size_code
                        if loads[size_code][i] != '':

                            # We save the number of item per crate for this size_code
                            index = list(grouped_dataframe['MODEL']).index(size_code)
                            nb_per_crate = grouped_dataframe['NBR_PER_CRATE'][index]

                            # For every unit of this crate on the load
                            for unit in range(int((loads[size_code][i])/nb_per_crate)):

                                # We pick a random SKU linked with the size_code
                                sku = self.tracker[size_code].random_pick(nb_per_crate)
                                data.append([load_number, category, load_length, sku, nb_per_crate, size_code])

            columns_title = [ws.cell(row=1, column=i).value for i in range(1, nbr_of_cols+1)]
            dataframe = pd.DataFrame(data=data, columns=columns_title)

            # We do a groupby to sum quantity column for the same SKU on the same load
            dataframe = dataframe.groupby(by=[column for column in columns_title if column != 'QUANTITY']).sum().reset_index()
            dataframe = dataframe[columns_title]

            # We push every line of data in the appropriate worksheet
            for i in dataframe.index:
                ws.append(list(dataframe.iloc[i].values))

            # Shape data as a dynamic table
            create_excel_table(ws, "APPROVED", columns_title)

    def write_unused_crates(self, ws, nbr_of_cols):
        """
        Writes the results for the unused crates in the worksheet passed as parameter
        :param ws: worksheet on which we write the results
        :param nbr_of_cols: number of columns in the worksheet
        """
        for model, sku_container in self.tracker.items():
            for sku, qty in sku_container.skus_dict.items():
                ws.append([sku, qty])

        # We save the columns titles
        columns_title = [ws.cell(row=1, column=i).value for i in range(1, nbr_of_cols + 1)]

        # We shape data as a dynamic table
        create_excel_table(ws, "UNUSED", columns_title)

    @staticmethod
    def positive(a):
        """
        Look if the input is positive, we don't check if it's an integer because it's automatically converted as an int
        in save_skus_and_quantitites and save_trailer_quantities.
        :param a: variable
        :return: raise an Exception if the input is not positive nor integer
        """
        raise NotImplementedError

    @staticmethod
    def get_complete_dataframe(skus_list, qty_list, crate_type):
        """
        Retrieves size_code and dimensions associated with our SKUs
        plus : 'NBR_PER_CRATE', 'STACK_LIMIT' and 'OVERHANG'

        :param skus_list: list of SKUs used in sql query to get the data
        :param qty_list: list of quantities associated to each SKU
        :param crate_type: 'W' for wood, 'M' for metal
        :return: pandas dataframe
        """
        # Initialization of column names for data that will be sent to LoadBuilder
        columns = ['QTY', 'SKU', 'MODEL', 'LENGTH', 'WIDTH', 'HEIGHT', 'NBR_PER_CRATE',
                   'CRATE_TYPE', 'STACK_LIMIT', 'OVERHANG']

        # Connection to SQL database that contains data needed
        sql_connect = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_0_MD_D_MATERIAL')

        # We look length of SKUs list and adapt the code
        if len(skus_list) == 1:
            end_of_query = """WHERE CRATE_SIZE.Material_Number = """ + "'" + str(skus_list[0]) + "'"

        else:
            end_of_query = """WHERE CRATE_SIZE.Material_Number in """ + str(tuple(skus_list))

        if crate_type == 'W':
            subquery = """(
            SELECT 
            M1.MATERIAL_NUMBER,
            M1.SIZE_DIMENSIONS,
            CEILING(MAX(M2.LENGTH)) AS LENGTH,
            CEILING(MAX(M2.WIDTH)) AS WIDTH,
            CEILING(MAX(M2.HEIGHT)) AS HEIGHT
			FROM MASTERDATA.DBO.MD_MARA AS M1 LEFT JOIN MASTERDATA.DBO.MD_MARA AS M2
			ON M1.REF_MAT_PACKED_IN_SAME_WAY = M2.MATERIAL_NUMBER
			GROUP BY M1.MATERIAL_NUMBER, M1.SIZE_DIMENSIONS) AS CRATE_SIZE
            """
            size_dimensions = """,RTRIM(a.SIZE_DIMENSIONS) AS SIZE_DIMENSIONS"""

            overhang_case = ""
        else:
            subquery = """(
            SELECT CLASSIF.MATERIAL_NUMBER, CLASSIF.CHARACTERISTIC_VALUE, HEIGHT,LENGTH,WIDTH, 'CRATE' AS CRATE_TYPE
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF LEFT JOIN MASTERDATA.DBO.MD_MARA AS MARA
            ON CLASSIF.CHARACTERISTIC_VALUE = MARA.MATERIAL_NUMBER LEFT JOIN ( SELECT *
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF
            WHERE CLASS = 'RETURNABLE_CRATE' AND LANGUAGE_CODE = 'E' AND CHARACTERISTIC_NAME = 'TEMPORARY_CRATE' AND CHARACTERISTIC_VALUE = 'Y') AS SKID
            ON CLASSIF.MATERIAL_NUMBER = SKID.MATERIAL_NUMBER
            WHERE CLASSIF.CLASS = 'RETURNABLE_CRATE' AND CLASSIF.LANGUAGE_CODE = 'E' AND CLASSIF.CHARACTERISTIC_NAME = 'CRATE_NUMBER' AND SKID.MATERIAL_NUMBER IS NULL
            
            UNION
            
            SELECT SKID.MATERIAL_NUMBER,ISNULL(C.CHARACTERISTIC_VALUE,999) AS CHARACTERISTIC_VALUE, ISNULL(H.CHARACTERISTIC_VALUE,999) AS HEIGHT ,ISNULL(L.CHARACTERISTIC_VALUE,999) AS LENGTH, ISNULL(W.CHARACTERISTIC_VALUE,999) AS WIDTH, 'SKID' AS CRATE_TYPE
            FROM (SELECT *
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF
            WHERE CLASS = 'RETURNABLE_CRATE' AND LANGUAGE_CODE = 'E' AND CHARACTERISTIC_NAME = 'TEMPORARY_CRATE' AND CHARACTERISTIC_VALUE = 'Y') AS SKID
            LEFT JOIN (SELECT *
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF
            WHERE CLASS = 'RETURNABLE_CRATE' AND LANGUAGE_CODE = 'E' AND CHARACTERISTIC_NAME = 'VEHIC_LENGHT_LEG1') AS L
            ON SKID.MATERIAL_NUMBER = L.MATERIAL_NUMBER LEFT JOIN (SELECT *
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF
            WHERE CLASS = 'RETURNABLE_CRATE' AND LANGUAGE_CODE = 'E' AND CHARACTERISTIC_NAME = 'VEHIC_WIDTH_LEG1') AS W
            ON SKID.MATERIAL_NUMBER = W.MATERIAL_NUMBER LEFT JOIN (SELECT *
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF
            WHERE CLASS = 'RETURNABLE_CRATE' AND LANGUAGE_CODE = 'E' AND CHARACTERISTIC_NAME = 'MIN_REQUIRED_LEG1') AS H
            ON SKID.MATERIAL_NUMBER = H.MATERIAL_NUMBER LEFT JOIN (SELECT *
            FROM  [MASTERDATA].[DBO].[MD_MATERIAL_SAP_CLASSIFICATION_VIEW] AS CLASSIF
            WHERE CLASS = 'RETURNABLE_CRATE' AND LANGUAGE_CODE = 'E' AND CHARACTERISTIC_NAME = 'CRATE_NUMBER') AS C
            ON SKID.MATERIAL_NUMBER = C.MATERIAL_NUMBER
            WHERE SKID.MATERIAL_NUMBER IS NOT NULL)
			as CRATE_SIZE"""

            size_dimensions = """,CASE WHEN MKT_MODEL_YEAR >= '2020' THEN LEFT(a.[SIZE_DIMENSIONS],1) ELSE a.SIZE_DIMENSIONS END AS SIZE_DIMENSIONS"""

            overhang_case = """WHEN MKT_MODEL_YEAR >= '2020'  THEN 0"""

        # Writing of our query
        sql_query = """ SELECT RTRIM(a.Material_Number) """ + size_dimensions + """
        ,CONVERT(int, CEILING(CRATE_SIZE.Length))
        ,CONVERT(int, CEILING(CRATE_SIZE.Width))
        ,CONVERT(int, CEILING(CRATE_SIZE.HEIGHT))
        ,CASE WHEN c.SUB_DIVISION = 'RYKER' THEN 2 ELSE 1 END
        ,CASE WHEN CRATE_TYPE = 'SKID' THEN 1 WHEN CRATE_SIZE.HEIGHT = 0 THEN 1 ELSE CONVERT(int, FLOOR(105/CRATE_SIZE.HEIGHT)) END
        ,CASE WHEN (CASE WHEN CRATE_TYPE = 'SKID' THEN 1 WHEN CRATE_SIZE.HEIGHT = 0 THEN 1 ELSE FLOOR(105/CRATE_SIZE.HEIGHT) END) = 1 THEN 0
        """ + overhang_case + """ ELSE 1 END
        FROM OTD_0_MD_D_MATERIAL as c LEFT JOIN masterdata.dbo.MD_MARA as a ON c.Material_number = a.Material_Number 
        LEFT JOIN """ + subquery + """ on c.Material_number = CRATE_SIZE.Material_Number
        LEFT JOIN [dbo].[OTD_1_P2P_F_PARAMETERS_CRATE_SKID] as SKID
        on a.Size_Dimensions = SKID.[CRATE_SIZE_SKID] """ + end_of_query

        print(sql_query)

        # Retrieve the data
        data = sql_connect.GetSQLData(sql_query)

        # Add each qty at the beginning of the good line of data and crate type after the stack limit
        for line in data:
            index_of_qty = skus_list.index(line[0])
            line.insert(0, qty_list[index_of_qty])
            line.insert(7, crate_type)

        return pd.DataFrame(data=data, columns=columns)

    @staticmethod
    def split_dataframes(complete_dataframe):
        """
        Takes the complete dataframe and split it in two
        First : [QTY | SKU |  MODEL (SIZE_CODE)] to keep track of link between SKUs and size_code
        Second : [QTY | MODEL | LENGTH | WIDTH | HEIGHT | NUMBER_PER_CRATE | CRATE_TYPE | STACK_LIMIT | OVERHANG ]
        The second will be group by MODEL

        :return: two pandas data frames
        """

        # We extract both dataframes needed by making copy of some parts on complete dataframe
        first_df = complete_dataframe[['QTY', 'SKU', 'MODEL']].copy()
        second_df = complete_dataframe[['QTY', 'MODEL', 'LENGTH', 'WIDTH', 'HEIGHT', 'NBR_PER_CRATE', 'CRATE_TYPE',
                                       'STACK_LIMIT', 'OVERHANG']].copy()

        # Do a groupby on second dataframe
        second_df = second_df.groupby(['MODEL', 'LENGTH', 'WIDTH', 'HEIGHT', 'CRATE_TYPE',
                                       'NBR_PER_CRATE', 'STACK_LIMIT', 'OVERHANG']).sum().reset_index()

        return first_df, second_df

    @staticmethod
    def tracker_initialization(sku_dataframe):
        """
        Initializes a dictionary of size_code (key) with SKUsContainer as value
        :param sku_dataframe: pandas dataframe containing [QTY | SKU |  MODEL (SIZE_CODE)]
        :return: dictionary ("tracker")
        """
        # We count the number of unique model
        models = set(sku_dataframe['MODEL'])

        # We intialize our tracker dict and put empty SKUsContainer in it
        tracker = {}
        for model in models:
            tracker[model] = SKUsContainer()

        # We fill every SKUsContainer
        for i in sku_dataframe.index:
            tracker[sku_dataframe['MODEL'][i]].add_sku(sku_dataframe['SKU'][i], sku_dataframe['QTY'][i])

        return tracker


class SKUsContainer:

    def __init__(self):
        self.skus_dict = {}

    def add_sku(self, sku, qty):
        self.skus_dict[sku] = qty

    def __remove_sku(self, sku):
        self.skus_dict.pop(sku)

    def random_pick(self, qty):
        """
        Randomly pick a SKU in the SKUsContainer and decrease its quantity by "qty"
        :param qty: (int) Number of times we pick the randomly chosen SKU
        :return: (str) SKU picked
        """

        # We pick a SKU randomly in our dictionary
        if len(self.skus_dict) - 1 == 0:
            rand_index = 0
        else:
            rand_index = randint(0, len(self.skus_dict) - 1)

        skus = list(self.skus_dict.keys())
        rand_sku = skus[rand_index]

        # We decrease its quantity by the quantity of this SKU in one crate
        self.skus_dict[rand_sku] -= qty

        # We remove the sku from our dictionary if the quantity is now null
        if self.skus_dict[rand_sku] == 0:
            self.__remove_sku(rand_sku)

        # We return the SKU
        return rand_sku

    def __repr__(self):
        return str(self.skus_dict)


def open_fastloads_box():

    set_project_name('ADHOC')
    root = Tk()
    fastloadsbox = FastLoadsBox(root)
    root.mainloop()


def build_dataframe(ws, fast_loads_input=False):

    """
    Build a data frame (used to store models' data)

    :param ws: Worksheet or part of the worksheet use to build de pandas dataframe
    :param fast_loads_input: bool indicating if the dataframe is built for the fast loads
    :return: Pandas dataframe

    """

    data_rows = []
    for row in ws:
        data_cols = []
        for cell in row:
            if cell.value is not None:
                data_cols.append(cell.value)
        if len(data_cols) == len(row):
            data_rows.append(data_cols)

    df = pd.DataFrame(data=data_rows[1:], columns=data_rows[0])

    if fast_loads_input:
        df = df.groupby(['SKU']).sum().reset_index()
        df = df[data_rows[0]]

    return df

