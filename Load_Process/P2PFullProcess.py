"""

Authors : Olivier Lefebvre
          Nicolas Raymond

This file creates loads for P2P

Last update : 2020-01-06
By : Nicolas Raymond

"""

import sys
from ParametersBox import OpenParameters, MissingP2PBox
from P2PFunctions import *
from ProcessValidation import validate_process
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from datetime import datetime, timedelta
import os

""" Used SQL Table and actions: """
# [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS_EMAIL_ADDRESS] : Get data, alter, delete, send
# [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS] : Get data, alter, delete, send
# [Business_Planning].[dbo].[OTD_1_P2P_F_PRIORITY_WITHOUT_INVENTORY] : Get data
# [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY] : Get data
# OTD_1_P2P_D_INCLUDED_INVENTORY : Get data
# OTD_1_P2P_F_HISTORICAL : Send data
# OTD_1_P2P_F_TRUCK_PARAMETERS : Get data

# Folder where the excel workbook is saved
saveFolder = 'S:\Shared\Business_Planning\Tool\Plant_to_plant\Optimus\P2P_excel_output\P2P_summary\\'

# Parameters variables
dayTodayComplete = pd.datetime.now().replace(second=0, microsecond=0)  # date to set in SQL for import date
dayToday = weekdays(0)  # Date to display in report
drybox_sanity_check = True  # Activates drybox load validation
printLoads = False  # Print created loads
save_log_file = True  # Save log file of process results
good_credit_for_max = False  # If set to true, allow only wishes with good credit to be used to satisfy max
MinWarning = False  # Add yellow filling as warning when minimum is not satisfied for a p2p
AutomaticRun = False  # set to True to automate code
validation = False     # set to True to validate the results received after the process
result_time_stamp = time_now_string()
general_folder = saveFolder+'P2P_Summary_'+dayToday+'\\'
result_folder = general_folder+result_time_stamp+'\\'
result_file = 'P2P_Summary_'+result_time_stamp  # Name of excel file with today's date
history_expiration_date = dayTodayComplete - timedelta(days=365)  # Expiration date of history data set one year ago


def p2p_full_process():

    """
    Executes P2P full process

    :return: summary of the full process in at the 'saveFolder' directory

    """

    if not AutomaticRun:
        if not OpenParameters():  # If user cancel request
            sys.exit()

    timeSinceLastCall('', False)

    ####################################################################################################################
    #                                                Get SQL DATA
    ####################################################################################################################

    # If SQL queries crash
    downloaded = False
    nbr_of_try = 0
    while not downloaded and nbr_of_try < 3:  # 3 trials, SQL Queries sometime crash for no reason
        nbr_of_try += 1
        try:
            downloaded = True

            # Get details on point from and shipping point
            shipping_points_names = shipping_point_names()

            # Get email addresses
            emails_list = get_emails_list('P2P')

            # Get plant to plant (parameters)
            p2ps_list, p2p_connection = get_parameter_grid()

            # Get plant to plant orders in which they should be written in the output
            p2ps_order_list = get_p2p_order(p2p_connection)

            # Get the wishlist
            wishlist = get_wish_list()

            # We get the nested shipping points (needs to be done before inventory recuperation)
            global DATAInclude
            get_nested_source_points(DATAInclude)

            # Get the inventory
            inventory = get_inventory_and_qa()

            #  Look if all point_from + shipping_point are in parameters
            missing_p2ps_list = get_missing_p2p()

        except pyodbc.Error as err:
            downloaded = False
            sql_state = err.args[1]
            print('SQL Query failed :', sql_state)

    # If SQL Queries failed
    if not downloaded:
        try:
            send_email(emails_list, result_file, 'SQL QUERIES FAILED')
        except:
            pass
        sys.exit()

    timeSinceLastCall('Get SQL DATA')

    # If there are missing P2P in parameters table
    if not AutomaticRun:
        if missing_p2ps_list:
            if not MissingP2PBox(missing_p2ps_list):
                sys.exit()

    timeSinceLastCall('', False)

    # Creation of results folders
    create_directory(general_folder)
    create_directory(result_folder)

    # Creation of log file
    create_log_file(result_folder, save_log_file)

    # Application of credit rule to satisfy max
    set_good_credit_for_max(good_credit_for_max)

    ####################################################################################################################
    #                                                 Excel Workbook declaration
    ####################################################################################################################

    # Initialization of workbook and filling option to warn user in output
    wb = Workbook()
    Warning_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

    # Summary
    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    summary_columns = ['POINT_FROM_NUMBER', 'POINT_FROM_NAME', 'SHIPPING_POINT_NUMBER',
                       'SHIPPING_POINT_NAME', 'NUMBER_OF_LOADS', 'NUMBER_OF_VEHICLES']
    worksheet_formatting(summary_ws, summary_columns, [30]*4 + [25, 25])

    # Approved loads
    approved_ws = wb.create_sheet("APPROVED")
    approved_columns = ['POINT_FROM', 'SHIPPING_POINT', 'LOAD_NUMBER', 'CATEGORY', 'LOAD_LENGTH',
                        'MATERIAL_NUMBER', 'QUANTITY', 'SIZE_DIMENSIONS']
    columns_width = [20]*len(approved_columns)
    worksheet_formatting(approved_ws, approved_columns, columns_width)

    # Unbooked
    unbooked_ws = wb.create_sheet("UNBOOKED")
    unbooked_columns = ['POINT_FROM', 'MATERIAL_NUMBER', 'QUANTITY']
    worksheet_formatting(unbooked_ws, unbooked_columns, [20, 20, 12])

    # Booked unused
    unused_ws = wb.create_sheet("BOOKED_UNUSED")  # The formatting will be done later

    # Approved Summarized
    sap_input_ws = wb.create_sheet("SAP INPUT")
    sap_input_columns = ['POINT_FROM', 'SHIPPING_POINT', 'MATERIAL_NUMBER', 'QUANTITY']
    worksheet_formatting(sap_input_ws, sap_input_columns, [20]*len(sap_input_columns))

    ####################################################################################################################
    #                                            Isolate perfect match
    ####################################################################################################################

    # We initialize a list that will contain all wish approved
    approved_wishes = find_perfect_match(wishlist, inventory, p2ps_list)

    ####################################################################################################################
    #                                    Set some LoadBuilder class' attributes
    ####################################################################################################################

    # We first set flatbed_48 as trailer reference of LoadBuilder for sanity check
    set_trailer_reference(get_trailers_data(['FLATBED_48'], [1]))

    # We set the attribute "validate_with_ref" for LoadBuilder class
    LoadBuilder.validate_with_ref = drybox_sanity_check

    ####################################################################################################################
    #                                                Create Loads
    ####################################################################################################################

    perfect_match_loads_construction(p2ps_list, approved_wishes, print_loads=printLoads)

    ####################################################################################################################
    #                             Try to Make the minimum number of loads for each P2P
    ####################################################################################################################

    satisfy_max_or_min(wishlist, inventory, p2ps_list, print_loads=printLoads)

    ####################################################################################################################
    #                             Try to Make the maximum number of loads for each P2P
    ####################################################################################################################

    satisfy_max_or_min(wishlist, inventory, p2ps_list, satisfy_min=False, print_loads=printLoads)

    ####################################################################################################################
    #                                       Distribution of leftover crates
    ####################################################################################################################

    distribute_leftovers(wishlist, inventory, p2ps_list)

    ####################################################################################################################
    #                                           Writing of the results
    ####################################################################################################################

    # We display loads create in each p2p for our own purpose and save load pictures in the results folder
    print('\n\nResults')
    for param in p2ps_list:
        print('\n\n')
        print(param.POINT_FROM, ' _ ', param.POINT_TO)
        print(len(param.LoadBuilder))
        print(param.LoadBuilder.trailers_done)
        print(param.LoadBuilder.get_loading_summary())  # Trailer are now sorted correctly

    # Initialization of a list to keep all data needed for the "APPROVED" ws and summary version ouptput for SAP
    approved_ws_data, sap_input_data = [], []

    line_index = 2  # To display a warning if number of loads is lower than parameters min

    # We get the connection to the history table while deleting expired data
    connection = clean_p2p_history(history_expiration_date)

    # We start to write the results
    for order in p2ps_order_list:  # To send data in excel workbook, order by : -point_from , -point_to

        # For the p2p that match with the order
        for param in [p2p for p2p in p2ps_list if (p2p.POINT_FROM == order[0] and p2p.POINT_TO == order[1])]:

            # We write a line in the summary worksheet
            summary_ws.append([param.POINT_FROM, shipping_points_names[param.POINT_FROM],
                               param.POINT_TO, shipping_points_names[param.POINT_TO],
                               len(param.LoadBuilder), param.get_nb_of_units()])

            # If the minimum is not fulfilled we warn the user with a different background color in the output
            if MinWarning and len(param.LoadBuilder) < param.LOADMIN:
                summary_ws.cell(row=line_index, column=len(summary_columns)).fill = Warning_fill

            # We update line index
            line_index += 1

            # If some loads were created
            if len(param.LoadBuilder) > 0:

                # We save p2p results to sql and in some lists for excel output
                param.save_full_process_results(connection, approved_ws_data, sap_input_data,
                                                dayTodayComplete, result_folder)

    # We compute booked unused
    possible_plant_to = compute_booked_unused(wishlist, inventory, p2ps_list)

    # We format BOOKED UNUSED worksheet
    unused_columns = ['POINT_FROM', 'MATERIAL_NUMBER', 'SIZE_DIMENSIONS', 'QUANTITY'] + possible_plant_to
    worksheet_formatting(unused_ws, unused_columns, [20, 25, 20, 15] + [15]*len(possible_plant_to))

    # We send inventory in unused_ws and unbooked_ws
    for inv in inventory:
        if inv.unused > 0:
            unused_ws.append(inv.lineToXlsx(possible_plant_to=possible_plant_to))
        if inv.QUANTITY - inv.unused > 0:
            unbooked_ws.append(inv.lineToXlsx(booked_unused=False))

    # We group by the APPROVED input data
    approved_frame = group_by_all_except_qty(approved_ws_data, approved_columns)

    # We send it in the output and save rows index on which we want to apply the grey filling
    last_point_from, last_shipping_point = approved_frame['POINT_FROM'][0], approved_frame['SHIPPING_POINT'][0]
    total_load_number, last_load_number = 1, 1
    indexes_to_fill = []

    for index, row in approved_frame[approved_columns].iterrows():

        # If we're looking at a different load
        if row.POINT_FROM != last_point_from or row.SHIPPING_POINT != last_shipping_point or\
                row.LOAD_NUMBER != last_load_number:

            # We update last load details and update the total_load_number
            last_point_from, last_shipping_point, last_load_number = row.POINT_FROM, row.SHIPPING_POINT, row.LOAD_NUMBER
            total_load_number += 1

        # We add the row index in indexes to fill its total_load_number is pair
        if (total_load_number % 2) == 0:
            indexes_to_fill.append(index + 2)

        approved_ws.append(list(row))

    # We filled rows that must be filled in the "APPROVED" worksheet
    apply_filling_to_rows(approved_ws, indexes_to_fill, len(approved_columns))

    # We group by the SAP input data and send it in the output
    sap_input_frame = group_by_all_except_qty(sap_input_data, sap_input_columns)

    for index, row in sap_input_frame[sap_input_columns].iterrows():
        sap_input_ws.append(list(row))

    # We create tables for the data of every worksheet
    create_excel_table(summary_ws, "Summary", summary_columns)
    create_excel_table(approved_ws, "Approved", approved_columns)
    create_excel_table(unbooked_ws, "Unbooked", unbooked_columns)
    create_excel_table(unused_ws, "Booked_unused", unused_columns)
    create_excel_table(sap_input_ws, "SAP_input", sap_input_columns)

    # We save the workbook and the reference
    reference = [savexlsxFile(wb, result_folder, result_file, Time=False)]

    # We send the emails
    send_email(emails_list, result_file, '', reference)

    # We validate the process' results if the user wants to
    if validation:
        validate_process(reference[0], p2ps_list, residuals_counter)

    # We open excel workbook
    os.system('start "excel" "'+str(reference[0])+'"')


if __name__ == '__main__':
    p2p_full_process()
